# -*- coding: utf-8 -*-
# @Time    : 2019/12/19 11:19
# @Author  : xianming yu
# @File    : 2python操控EXCEL之读取数据后处理实例.py

"""
把修改好的列的值,存放到新列当中
"""

import openpyxl

wb = openpyxl.load_workbook("example.xlsx")  # 打开工作簿

wb_names = wb.sheetnames  # 获取example.xlsx工作簿下的表单名称
sheet1 = wb["Sheet1"]  # 操作example.xlsx工作簿下的Sheet1表单
list_a = []  # 创建一个空列表
col_range = sheet1["E:E"]  # 取Sheet1表单中E:E的值赋值给到的col_range变量
for col in col_range:  # 遍历E:E列的值
    list_a.append(col.value)  # 把E的值取出来以后放到list_a列表当中
list_b = []  # 创建一个空列表
for i in list_a:  # 遍历list_a列表
    list_b.append(i[:4])  # 取下标0-4值,并添加到list_b列表当中
for i in range(len(list_b)):  # 遍历list_b列表值的下标
    sheet1.cell(i + 1, 2).value = list_b[i]  # 写入excle表格  i+1代表行,2代表列,诠释:从第2列第1个单元格开始写入
wb.save("1.xlsx")  # 把修改好的数据保存到一个为1.xlsx表格当中


"""
excel表格打印输出对齐
"""
