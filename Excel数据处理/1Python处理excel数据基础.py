# -*- coding: utf-8 -*-
# @Time    : 2019-12-18 9:02
# @Author  : xianming yu
# @File    : Python处理excel数据.py
"""学习纪要"""
"""
安装:openpyxl
"""
"""
处理对象:
    1.工作簿(workbook)
    2.表单(worksheet) 如sheet1,sheet2等等
    3.行,列,单元格(row,column,cell)
"""


import openpyxl

#打开工作簿
wb = openpyxl.load_workbook("C:\\Users\Administrator\Desktop\\temp\\example.xlsx") #打开excel工作簿类型文件,可以带上路径
"""
#获取表单名称
a=wb.sheetnames  #获取example.xlsx工作簿下的表单名称
print(a)         # 打印表单名称,返回的类型是一个列表(list)
#遍历表单名称
for i in a:
    print(i.title()) # 遍历表单名称,返回的类型是字符串(str)
#创建表单
mysheet= wb.create_sheet("mysheet")  # 在example.xlsx工作簿下创建一个表单
print(wb.sheetnames)  # 再一次打印表单名称看是否创建表单成功

#操作表单1
sheet2 = wb["Sheet2"]  # 操作example.xlsx工作簿下的sheet2表单
print(sheet2["A1"])  #打印单元格对象
print(sheet2["A1"].value) #打印单元格中的值


#操作表单2
ws=wb.active  #active这个函数默认操作是example.xlsx工作簿中的第1个表单
print(ws) 
print(ws["A1"])  #打印单元格对象
print(ws["A2"].value)  #打印单元格中的值


#在Sheet1的表单中取行和列的值
sheet1=wb["Sheet1"]
c=sheet1["B1"]
print("Row {},Column {} is {}".format(c.row,c.column,c.value)) #输出结果: Row 1,Column 2 is 供应商
print("Cell {} is {}".format(c.coordinate,c.value)) #输出结果: Cell B1 is 供应商

#在Sheet2的表单中取行和列的值
ws=wb.active
print(ws.cell(row=1,column=2)) # 取第1行第2列的对象   输出结果:<Cell 'Sheet2'.B1>
print(ws.cell(row=1,column=2).value) # 取第1行第2列的值   输出结果:供应商
for i in range(1,8,2):
    print(i,ws.cell(row=2,column=2).value)
"""

#在表单中取行和列
"""
#行操作
sheet1=wb.active
row6 = sheet1[6] #表示取表单中第6行的数据
print(row6) #打印表单中第6行的数据
for  row in row6: #遍历第6行数据的值
    print(row.value,end=' ')
    
row_range=sheet1[2:6] #取表单中第2行到第6行的切片
print(row_range)  #打印表单中第2行到第6行的切片
#第一种行遍历方法:
for row in row_range: # 遍历表单中第2行到第6行切片的值
    for row_value in row:
        print(row_value.value,end=" ")
    print()

#第二种行遍历方法
for row in sheet1.iter_rows(min_row=1,max_col=3): #表示取3列中每一行的数据,也就是一个数据区间
    for cell in row:
        print(cell.value,end=" ")
    print()


#列操作
colc=sheet1["C"]  #取表单中的C列
print(colc) #打印C列
print(colc[2].value) #取表单中C列下第3个单元格的值,colc[2]中括号中的2代表的是3,因为下标是从零开始
col_range=sheet1["B:C"] #取表单中B列到C列的切片
print(col_range) #打印B列到C列的切片
for col in col_range: # 遍历表单中B列到C列切片的值
    for cell in col:
        print(cell.value)
    print()

#单元格切片
cell_range=sheet1["A2:E8"]
for row in cell_range:
    for cellobj in row:
        print(cellobj.coordinate,cellobj.value,end=" ")
    print()
"""
#统计表单中行和列
"""
sheet2=wb["Sheet2"]
print("{}*{}".format(sheet2.max_row,sheet2.max_column))
"""


#列名称字母与数字之间的转换
"""
from openpyxl.utils import get_column_letter,column_index_from_string
#其中get_column_letter是把列名称由数字转换成字母,column_index_from_string是把列名称由字母转换成数字
print(get_column_letter(2),get_column_letter(47),get_column_letter(110)) #对应的是B列,AU列,DF列
print(column_index_from_string("B"),column_index_from_string("D")) #对应的是第2列和第4列
"""



