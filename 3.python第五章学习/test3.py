
""""
处理excel表格数据

import openpyxl
from openpyxl.styles import Alignment

wb = openpyxl.load_workbook("example.xlsx")  # 打开工作簿
sheet1=wb["Sheet1"]

print("{}*{}".format(sheet1.max_row,sheet1.max_column))

for row in sheet1.iter_rows(min_row=1,max_col=24): #表示取3列中每一行的数据,也就是一个数据区间
    for cell in row:
        cell_new=str(cell.value)
        print(format(cell_new, "19s"), end=" ")
        # print(cell.value,end=" ")
    print()
"""
"""
文件清理定时任务


from threading import Timer
import datetime
# 每隔两秒执行一次任务
def printHello():
    print('TimeNow:%s' % (datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
    t = Timer(10, printHello)
    t.start()

if __name__ == "__main__":
    printHello()
"""

"""
import  os
import time
import datetime
path=input("需要清除文件的目录:")
sele_path=os.listdir(path)
for i in sele_path:
    a=path+"\\"+i
    b=time.localtime(os.stat(a).st_ctime)
    b=time.strftime("%Y%m%d",b)
    # print(b)
    today = datetime.datetime.now()
    offset = datetime.timedelta(days=-3)
    re_date = (today + offset).strftime('%Y%m%d')
    if re_date>b:
        os.remove(a)
    else:
        pass
    # print(i,time.ctime(os.stat(path+"\\"+i).st_ctime) )#文件的创建时间
"""

"""
import datetime
today = datetime.datetime.now()
offset = datetime.timedelta(days=3)
re_date = (today + offset).strftime('%Y%m%d')
today=today.strftime('%Y%m%d')
print("re_date:",re_date)
print(today)
print(re_date)
print(today>re_date)
"""

"""
import turtle
for y in range(0,200,20):
    for x in range(0,200,20):
        turtle.goto(x,y)
        turtle.pendown()
    turtle.penup()
for x in range(0,200,20):
    for y in range(0,200,20):
        turtle.goto(x,y)
        turtle.pendown()
    turtle.penup()
turtle.done()
"""



"""
for i in range(0, 300, 40):
    for j in range(0, 400, 100):
        turtle.goto(j, i)
        turtle.pendown()
        turtle.write(str(j) + "," + str(i))
    turtle.penup()
turtle.done()
"""
"""
import turtle

turtle.showturtle()
turtle.goto(200, 100)
turtle.goto(-200, 100)
turtle.goto(-200, -100)
turtle.goto(200, -100)
turtle.goto(200, 100)
x, y = eval(input("输入x和y的坐标:"))
turtle.penup()
turtle.goto(x, y)
turtle.pendown()
turtle.dot(10, "black")
x=abs(x) #利用abs函数把x转换为绝对值
y=abs(y) #利用abs函数把y转换为绝对值
if x<200 and y<100:
    print("矩形内")
elif x==200 and y<100:
     print("矩形宽高边上")
elif x<200 and y==100:
    print("矩形长边上")
else:
    print("矩形外!")
turtle.done()
"""
import datetime
# a=datetime.date.today()
# print(a)
# print(a.year)
# print(a.month)
# print(a.day)
# b=datetime.date(2019,3,15)
# c=datetime.date(2020,4,13)
# print(b.__eq__(c))
# print(b.__eq__(c))
# print(b.__gt__(c))
# print(b.__le__(c))
# print(b.__lt__(c))
# print(b.__ne__(c))
# d=datetime.date.today()
# print(d)
# e=datetime.datetime.now()+datetime.timedelta(days=1)
# print(e.strftime('%Y-%m-%d'))
# print(d.__eq__(e))

#函数调试
# import pysnooper
# @pysnooper.snoop()
# def number(x,y):
#     for i in range(x,y):
#         j=0
#         while j<i:
#             print(j,end=" ")
#             j+=1
# number(1,5)

"""
中国象棋
import turtle
turtle.setup(width=800, height=880, startx=600, starty=0)
turtle.screensize(300, 300, "#D6C49E")
t=turtle.Pen()
t.speed(5)      #移动速度
t.pencolor()  #颜色
t.pensize(1.5)  #宽度
t.penup()      #抬起
t.goto(-240,270)   #起始点
t.pendown()      #下笔
for i in range(1,6,1):
    t.forward(480)  #画480px
    t.penup()    #抬起
    t.right(90)  #右转90度
    t.forward(60)  #移动60px
    t.right(90)    #右转90度
    t.pendown()   #下笔
    t.forward(480)  #画480px
    t.penup()   #抬起
    t.left(90)   #左转90度
    t.forward(60)  #移动60px
    t.left(90)   #左转90度
    t.pendown()   #下笔
t.left(90)
t.penup()
t.forward(60)
t.pendown()
for i in range(1,5,1):
    t.forward(240)
    t.penup()
    t.forward(60)
    t.pendown()
    t.forward(240)
    t.right(90)
    t.forward(60)
    t.right(90)
    t.forward(240)
    t.penup()
    t.forward(60)
    t.pendown()
    t.forward(240)
    t.left(90)
    t.forward(60)
    t.left(90)
t.forward(540)
t.left(90)
t.forward(480)
t.left(90)
t.forward(540)
t.left(90)
t.forward(180)
t.left(45)
t.forward(120*1.414)
t.left(45)
t.forward(-120)
t.left(45)
t.forward(120*1.414)
t.penup()
t.goto(-60,270)
t.pendown()
t.right(180)
t.forward(120*1.414)
t.right(45)
t.forward(-120)
t.right(45)
t.forward(120*1.414)
def angle(x,y):
    t.penup()
    t.goto(x+5,y+5)
    t.pendown()
    t.setheading(0)  #设置当前角度朝向
    t.forward(5)
    t.goto(x+5,y+5)
    t.left(90)
    t.forward(5)
    t.penup()
    t.goto(x+5,y-5)
    t.pendown()
    t.setheading(0)
    t.forward(5)
    t.goto(x+5,y-5)
    t.left(90)
    t.forward(-5)
    t.penup()
    t.goto(x-5,y+5)
    t.pendown()
    t.setheading(0)
    t.forward(-5)
    t.goto(x-5,y+5)
    t.left(90)
    t.forward(5)
    t.penup()
    t.goto(x-5,y-5)
    t.pendown()
    t.setheading(0)
    t.forward(-5)
    t.goto(x-5,y-5)
    t.left(90)
    t.forward(-5)

def a(x,y):
    t.penup()
    t.goto(x-5,y+5)
    t.pendown()
    t.setheading(0)
    t.forward(-5)
    t.goto(x-5,y+5)
    t.left(90)
    t.forward(5)
    t.penup()
    t.goto(x-5,y-5)
    t.pendown()
    t.setheading(0)
    t.forward(-5)
    t.goto(x-5,y-5)
    t.left(90)
    t.forward(-5)

def v(x,y):
    t.penup()
    t.goto(x+5,y+5)
    t.pendown()
    t.setheading(0)
    t.forward(5)
    t.goto(x+5,y+5)
    t.left(90)
    t.forward(5)
    t.penup()
    t.goto(x+5,y-5)
    t.pendown()
    t.setheading(0)
    t.forward(5)
    t.goto(x+5,y-5)
    t.left(90)
    t.forward(-5)
    t.penup()
angle(180,150)
angle(-180,150)
angle(180,-150)
angle(-180,-150)
angle(-120,90)
angle(0,90)
angle(120,90)
angle(-120,-90)
angle(0,-90)
angle(120,-90)

a(240,90)
a(240,-90)
v(-240,-90)
v(-240,90)
t.penup()
t.goto(-250,-280)
t.pendown()
t.pensize(8)
t.forward(560)
t.right(90)
t.forward(500)
t.right(90)
t.forward(560)
t.right(90)
t.forward(500)
t.right(90)
turtle.done()
"""
for  i in range(8):
    for j in range(8):



